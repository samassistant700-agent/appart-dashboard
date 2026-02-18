#!/usr/bin/env python3
"""
Script de conversion xlsx → JSON pour les appartements en mode LOCATION
"""

import json
import openpyxl
from datetime import datetime

def clean_phone_number(tel):
    """Nettoie le numéro de téléphone : 33467603160 → 04 67 60 31 60"""
    if not tel:
        return ""
    
    tel_str = str(tel).strip()
    
    # Enlever le "=" initial si présent (formule Excel)
    if tel_str.startswith("="):
        tel_str = tel_str[1:]
    
    # Garder seulement les chiffres
    tel_str = ''.join(c for c in tel_str if c.isdigit())
    
    # Si commence par "33", remplacer par "0"
    if tel_str.startswith("33") and len(tel_str) > 9:
        tel_str = "0" + tel_str[2:]
    
    # Formater en groupes de 2 chiffres : 04 67 60 31 60
    if len(tel_str) == 10:
        formatted = " ".join([tel_str[i:i+2] for i in range(0, 10, 2)])
        return formatted
    
    return tel_str

def parse_equipments(equipments_str):
    """Parse la colonne 'Parking / Cave / Terrasse / Clim'"""
    if not equipments_str:
        return {"parking": False, "cave": False, "terrasse": False, "clim": False}
    
    lower_str = str(equipments_str).lower()
    
    return {
        "parking": "parking" in lower_str,
        "cave": "cave" in lower_str,
        "terrasse": "terrasse" in lower_str,
        "clim": "clim" in lower_str or "climatisation" in lower_str
    }

def get_type_format(nb_pieces):
    """Génère le type format depuis Nb de pièces"""
    mapping = {
        1: "Studio",
        2: "T2",
        3: "T3",
        4: "T4"
    }
    return mapping.get(nb_pieces, "T5+")

def convert_xlsx_to_json(xlsx_path, output_path):
    """Convertit le fichier xlsx en JSON"""
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    # Lire les en-têtes
    headers = [cell.value for cell in ws[1]]
    
    # Créer un mapping colonne -> index
    col_index = {header: i for i, header in enumerate(headers)}
    
    appartements = []
    id_counter = 1
    
    # Parcourir les lignes (à partir de la ligne 2)
    for row in ws.iter_rows(min_row=2, values_only=True):
        quartier = row[col_index["Quartier"]]
        
        # Filtrer : garder seulement les lignes où "Quartier" n'est pas vide
        if not quartier or str(quartier).strip() == "":
            continue
        
        # Type (meublé)
        type_value = row[col_index["Type"]]
        meuble = str(type_value).strip() == "Meublé"
        
        # Prix (loyer mensuel)
        loyer = row[col_index["Prix"]]
        try:
            loyer = float(loyer) if loyer else 0
        except (ValueError, TypeError):
            loyer = 0
        
        # Surface
        surface = row[col_index["Surface au m²"]]
        try:
            surface = int(surface) if surface else 0
        except (ValueError, TypeError):
            surface = 0
        
        # Nb de pièces
        nb_pieces = row[col_index["Nb de pièces"]]
        try:
            nb_pieces = int(nb_pieces) if nb_pieces else 0
        except (ValueError, TypeError):
            nb_pieces = 0
        
        # DPE
        dpe = row[col_index["DPE"]]
        if dpe:
            dpe = str(dpe).strip().upper()
        else:
            dpe = ""
        
        # Chauffage
        chauffage = row[col_index["Type de chauffage"]]
        chauffage = str(chauffage).strip() if chauffage else ""
        
        # Charges annuelles
        charges_annuelles = row[col_index["Estimation des charges annuelles"]]
        try:
            charges_annuelles = float(charges_annuelles) if charges_annuelles else 0
        except (ValueError, TypeError):
            charges_annuelles = 0
        
        # Charges mensuelles
        charges_mensuelles = charges_annuelles / 12 if charges_annuelles > 0 else 0
        
        # Équipements
        equipments_str = row[col_index["Parking / Cave / Terrasse / Clim"]]
        equipements = parse_equipments(equipments_str)
        
        # État
        etat = row[col_index["État"]]
        etat = str(etat).strip() if etat else ""
        
        # Dates
        date_publication = row[col_index["Date de publication"]]
        if isinstance(date_publication, datetime):
            date_publication = date_publication.strftime("%Y-%m-%d")
        else:
            date_publication = str(date_publication).strip() if date_publication else ""
        
        date_contact = row[col_index["Date de prise de contact"]]
        if isinstance(date_contact, datetime):
            date_contact = date_contact.strftime("%Y-%m-%d")
        else:
            date_contact = str(date_contact).strip() if date_contact else ""
        
        date_visite = row[col_index["Rendez-vous visite"]]
        if isinstance(date_visite, datetime):
            date_visite = date_visite.strftime("%Y-%m-%d")
        else:
            date_visite = str(date_visite).strip() if date_visite else ""
        
        # Contact
        contact = row[col_index["Contact"]]
        contact = str(contact).strip() if contact else ""
        
        # Téléphone
        tel = row[col_index["Tel"]]
        tel = clean_phone_number(tel)
        
        # Adresse
        adresse = row[col_index["Adresse"]]
        adresse = str(adresse).strip() if adresse else ""
        
        # Site web
        site_web = row[col_index["Site web"]]
        site_web = str(site_web).strip() if site_web else ""
        
        # Notes
        notes = row[col_index["Notes"]]
        notes = str(notes).strip() if notes else ""
        
        # Construire l'objet JSON
        appartement = {
            "id": id_counter,
            "quartier": str(quartier).strip(),
            "type": get_type_format(nb_pieces),
            "meublé": meuble,
            "loyer": loyer,
            "charges": round(charges_mensuelles, 2),
            "charges_annuelles": round(charges_annuelles, 2),
            "surface": surface,
            "pieces": nb_pieces,
            "dpe": dpe,
            "chauffage": chauffage,
            "depotGarantie": loyer,  # 1 mois de loyer
            "parking": equipements["parking"],
            "cave": equipements["cave"],
            "terrasse": equipements["terrasse"],
            "clim": equipements["clim"],
            "ascenseur": False,  # Champ manquant
            "balcon": False,     # Champ manquant
            "etat": etat,
            "datePublication": date_publication,
            "dateContact": date_contact,
            "dateVisite": date_visite,
            "contact": contact,
            "tel": tel,
            "adresse": adresse,
            "siteWeb": site_web,
            "notes": notes
        }
        
        appartements.append(appartement)
        id_counter += 1
    
    # Écrire le JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(appartements, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Conversion terminée !")
    print(f"📊 {len(appartements)} appartements exportés")
    print(f"📁 Fichier JSON : {output_path}")
    
    return appartements

if __name__ == "__main__":
    xlsx_path = "/home/ubuntu/.openclaw/media/inbound/file_38---a0a33ffb-dfcc-46a4-ac24-82a5a21d5955.xlsx"
    output_path = "/home/ubuntu/.openclaw/workspace/appart-dashboard/data-location-real.json"
    
    appartements = convert_xlsx_to_json(xlsx_path, output_path)
